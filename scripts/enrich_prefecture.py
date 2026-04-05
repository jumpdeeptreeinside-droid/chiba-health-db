#!/usr/bin/env python3
"""
既存の都道府県DBに薬局機能フラグ + 2SFCAスコアを追加する汎用スクリプト

使い方:
  python3 enrich_prefecture.py --code 13 --shisetsu /path/to/shisetsu.xlsx

機能:
  1. 施設基準Excelから薬局機能フラグ9種を追加
  2. 全薬局 + 機能別の2SFCA分析
  3. 結果をDBに保存
"""

import argparse
import json
import math
import re
import sqlite3
import sys
import openpyxl
from collections import defaultdict
from pathlib import Path

PREF_NAMES = {
    1:"北海道",2:"青森県",3:"岩手県",4:"宮城県",5:"秋田県",6:"山形県",7:"福島県",
    8:"茨城県",9:"栃木県",10:"群馬県",11:"埼玉県",12:"千葉県",13:"東京都",14:"神奈川県",
    15:"新潟県",16:"富山県",17:"石川県",18:"福井県",19:"山梨県",20:"長野県",
    21:"岐阜県",22:"静岡県",23:"愛知県",24:"三重県",
    25:"滋賀県",26:"京都府",27:"大阪府",28:"兵庫県",29:"奈良県",30:"和歌山県",
    31:"鳥取県",32:"島根県",33:"岡山県",34:"広島県",35:"山口県",
    36:"徳島県",37:"香川県",38:"愛媛県",39:"高知県",
    40:"福岡県",41:"佐賀県",42:"長崎県",43:"熊本県",44:"大分県",45:"宮崎県",46:"鹿児島県",47:"沖縄県",
}

FUNCTION_FLAGS = {
    'kakaritsuke': ['かかりつけ薬剤師'],
    'chiiki_shien': ['地域支援体制加算'],
    'zaitaku_sogo': ['在宅薬学総合体制'],
    'mukin': ['無菌製剤処理'],
    'zaitaku_cv': ['在宅中心静脈栄養'],
    'zaitaku_mayaku': ['在宅患者医療用麻薬'],
    'medical_dx': ['医療ＤＸ推進', '医療DX推進'],
    'generic': ['後発医薬品調剤体制'],
    'renkei_kyoka': ['連携強化加算'],
}


def add_function_flags(conn, shisetsu_path):
    """施設基準Excelから薬局機能フラグを追加"""
    print("\n--- 薬局機能フラグ追加 ---")
    cur = conn.cursor()

    for flag in FUNCTION_FLAGS:
        try:
            cur.execute(f"ALTER TABLE pharmacies ADD COLUMN func_{flag} INTEGER DEFAULT 0")
        except:
            cur.execute(f"UPDATE pharmacies SET func_{flag}=0")
    conn.commit()

    # Excel解析: 名称(空白除去) → [届出名称リスト]
    wb = openpyxl.load_workbook(shisetsu_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    kikan_by_name = defaultdict(list)
    header_found = False
    col_name = 7       # default: 医療機関名称
    col_todokede = 13  # default: 受理届出名称
    for row in ws.iter_rows(values_only=True):
        if row[0] == '項番':
            # ヘッダーから列位置を動的に取得
            for i, v in enumerate(row):
                if v == '医療機関名称':
                    col_name = i
                if v == '受理届出名称':
                    col_todokede = i
            header_found = True
            continue
        if not header_found:
            continue
        name = str(row[col_name] or '').strip().replace('　', '').replace(' ', '')
        todokede = str(row[col_todokede] or '').strip()
        if name and todokede:
            kikan_by_name[name].append(todokede)
    wb.close()
    print(f"  Excel薬局: {len(kikan_by_name)}件")

    cur.execute('SELECT id, name FROM pharmacies')
    all_ph = cur.fetchall()
    matched = 0

    for pid, pname in all_ph:
        pn = str(pname).strip().replace('　', '').replace(' ', '')
        if pn not in kikan_by_name:
            continue
        matched += 1
        for todokede in kikan_by_name[pn]:
            for flag_name, patterns in FUNCTION_FLAGS.items():
                if any(pat in todokede for pat in patterns):
                    cur.execute(f"UPDATE pharmacies SET func_{flag_name}=1 WHERE id=?", (pid,))
    conn.commit()

    total = len(all_ph)
    print(f"  マッチ: {matched}/{total} ({matched/total*100:.1f}%)")
    for flag, pats in FUNCTION_FLAGS.items():
        cur.execute(f"SELECT COUNT(*) FROM pharmacies WHERE func_{flag}=1")
        cnt = cur.fetchone()[0]
        print(f"    {pats[0][:15]}: {cnt} ({cnt/total*100:.1f}%)")


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (math.sin(dlat/2)**2
         + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))
         * math.sin(dlon/2)**2)
    return R * 2 * math.asin(math.sqrt(a))


def gaussian_weight(d, t):
    if d > t: return 0
    b = t/2; return math.exp(-(d**2)/(b**2))


def run_2sfca(meshes, supplies, threshold):
    ratios = {}
    for s_id, s_lat, s_lon in supplies:
        wp = sum(gaussian_weight(haversine_km(s_lat, s_lon, m[1], m[2]), threshold) * m[3]
                 for m in meshes)
        ratios[s_id] = 1.0/wp if wp > 0 else 0
    access = {}
    for m in meshes:
        sc = sum(gaussian_weight(haversine_km(m[1], m[2], s[1], s[2]), threshold)
                 * ratios.get(s[0], 0) for s in supplies)
        access[m[0]] = sc
    return access


def compute_2sfca(conn):
    """全薬局 + 主要機能別の2SFCA"""
    print("\n--- 2SFCA分析 ---")
    cur = conn.cursor()

    cur.execute("SELECT mesh_code, lat, lon, population FROM population_mesh WHERE population >= 100 AND lat > 1")
    meshes = cur.fetchall()
    if not meshes:
        print("  メッシュデータなし、スキップ")
        return

    # 全薬局
    cur.execute("SELECT id, lat, lon FROM pharmacies WHERE lat > 1 AND lon > 1")
    all_sups = [(r[0], r[1], r[2]) for r in cur.fetchall()]
    if not all_sups:
        print("  有効座標の薬局なし、スキップ")
        return

    print(f"  メッシュ: {len(meshes)}, 薬局: {len(all_sups)}")

    # 全薬局2SFCA
    print("  全薬局...")
    access_all = run_2sfca(meshes, all_sups, 5.0)
    try:
        cur.execute("ALTER TABLE population_mesh ADD COLUMN access_phar_5km REAL")
    except:
        pass
    for mc, sc in access_all.items():
        cur.execute("UPDATE population_mesh SET access_phar_5km=? WHERE mesh_code=?", (sc, mc))
    conn.commit()

    # 機能別2SFCA（主要3機能のみ、速度のため）
    # 存在するカラムを確認
    cur.execute("PRAGMA table_info(pharmacies)")
    existing_cols = {r[1] for r in cur.fetchall()}

    for func_col, func_label in [
        ("func_kakaritsuke", "かかりつけ"),
        ("func_mukin", "無菌製剤"),
        ("func_zaitaku_sogo", "在宅総合"),
    ]:
        if func_col not in existing_cols:
            print(f"  {func_label}: カラム{func_col}なし、スキップ")
            continue
        cur.execute(f"SELECT id, lat, lon FROM pharmacies WHERE lat > 1 AND lon > 1 AND {func_col}=1")
        sups = [(r[0], r[1], r[2]) for r in cur.fetchall()]
        if not sups:
            continue
        print(f"  {func_label}: {len(sups)}件...")
        access = run_2sfca(meshes, sups, 5.0)
        col = f"access_{func_col}"
        try:
            cur.execute(f"ALTER TABLE population_mesh ADD COLUMN {col} REAL")
        except:
            pass
        for mc, sc in access.items():
            cur.execute(f"UPDATE population_mesh SET {col}=? WHERE mesh_code=?", (sc, mc))
    conn.commit()
    print("  完了")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--code", type=int, required=True)
    parser.add_argument("--shisetsu", type=str, help="施設基準Excelパス")
    args = parser.parse_args()

    pref_code = args.code
    pcode = f"{pref_code:02d}"
    pref_name = PREF_NAMES.get(pref_code, "不明")

    # DB探索
    candidates = [
        Path.home() / f"prefdb_{pcode}_{pref_name}" / f"{pcode}_iryo.db",
        Path.home() / "chiba_pdf_db" / "chiba_iryo.db" if pref_code == 12 else None,
        Path.home() / "osaka_pdf_db" / "osaka_iryo.db" if pref_code == 27 else None,
    ]
    db_path = None
    for c in candidates:
        if c and c.exists():
            db_path = c
            break

    if not db_path:
        print(f"ERROR: {pref_name}のDBが見つかりません")
        sys.exit(1)

    print(f"=== {pref_name} エンリッチ ===")
    print(f"DB: {db_path}")

    conn = sqlite3.connect(db_path)

    if args.shisetsu:
        shisetsu = Path(args.shisetsu)
        if shisetsu.exists():
            add_function_flags(conn, shisetsu)
        else:
            print(f"施設基準ファイル未発見: {shisetsu}")

    compute_2sfca(conn)
    conn.close()
    print(f"\n完了: {db_path}")


if __name__ == "__main__":
    main()
