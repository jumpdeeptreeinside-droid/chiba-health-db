"""
NDBオープンデータ（第8回・2020年度）から千葉県の疾患リスク指標を
disease_burden テーブルに追加するスクリプト

データソース: 厚労省NDBオープンデータ 特定健診検査結果（二次医療圏別）
https://www.mhlw.go.jp/stf/seisakunitsuite/bunya/0000177221_00012.html
"""

import sqlite3
import openpyxl
from pathlib import Path

BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "chiba_iryo.db"

# 千葉県の二次医療圏コードと名称
CHIBA_AREAS = {
    1201: "千葉",
    1202: "東葛南部",
    1203: "東葛北部",
    1204: "印旛",
    1205: "香取海匝",
    1206: "山武長生夷隅",
    1207: "安房",
    1208: "君津",
    1209: "市原",
}

# 各指標ファイルと疾患マッピング
# (ファイル名, 疾患名, 異常判定の閾値カテゴリリスト)
INDICATORS = [
    {
        "file": "ndb_hba1c_iryo.xlsx",
        "disease": "糖尿病（HbA1c高値）",
        "abnormal_categories": ["8.4以上", "8.0以上8.4未満", "6.5以上8.0未満"],
        "description": "HbA1c 6.5%以上（糖尿病域）",
    },
    {
        "file": "ndb_bp_iryo.xlsx",
        "disease": "高血圧（収縮期血圧高値）",
        "abnormal_categories": ["180以上", "160以上180未満", "140以上160未満"],
        "description": "収縮期血圧 140mmHg以上",
    },
    {
        "file": "ndb_ldl_iryo.xlsx",
        "disease": "脂質異常症（LDL高値）",
        "abnormal_categories": ["180以上", "160以上180未満", "140以上160未満"],
        "description": "LDLコレステロール 140mg/dl以上",
    },
    {
        "file": "ndb_bmi_iryo.xlsx",
        "disease": "肥満（BMI高値）",
        "abnormal_categories": [
            "40.0以上", "35.0以上40.0未満",
            "30.0以上35.0未満", "25.0以上30.0未満",
        ],
        "description": "BMI 25以上",
    },
    {
        "file": "ndb_fbs_iryo.xlsx",
        "disease": "糖尿病（空腹時血糖高値）",
        "abnormal_categories": ["126以上"],
        "description": "空腹時血糖 126mg/dl以上",
    },
]

DATA_SOURCE = "NDBオープンデータ第8回（2020年度）特定健診"
YEAR = 2020


def safe_int(val):
    """'‐'や'-'やNoneを0に変換"""
    if val is None or val == "‐" or val == "-" or val == "－":
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def extract_chiba_data(filepath, abnormal_categories):
    """
    Excelファイルから千葉県の二次医療圏別データを抽出
    Returns: {area_code: {"total": 総受診者数, "abnormal": 異常者数}}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    results = {}
    current_area_code = None
    in_chiba = False

    for row in ws.iter_rows(values_only=True):
        # 都道府県名で千葉県区間を検出
        if row[0] == "千葉県":
            in_chiba = True
        elif row[0] is not None and row[0] != "千葉県" and in_chiba:
            # 次の都道府県に入った → 千葉県区間終了
            break

        if not in_chiba:
            continue

        # 医療圏コード
        if row[1] is not None:
            current_area_code = row[1]
            if current_area_code not in CHIBA_AREAS:
                continue
            if current_area_code not in results:
                results[current_area_code] = {"total": 0, "abnormal": 0}

        if current_area_code not in CHIBA_AREAS:
            continue

        category = row[3]
        if category is None:
            continue

        # 男中計(col 11) + 女中計(col 19) = 合計人数
        male_total = safe_int(row[11])
        female_total = safe_int(row[19]) if len(row) > 19 else 0
        count = male_total + female_total

        results[current_area_code]["total"] += count

        if category in abnormal_categories:
            results[current_area_code]["abnormal"] += count

    wb.close()
    return results


def create_table(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS disease_burden (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            municipality TEXT,
            medical_area TEXT,
            disease_name TEXT,
            patient_count INTEGER,
            medical_cost REAL,
            year INTEGER,
            data_source TEXT
        )
    """)
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_disease_burden_area "
        "ON disease_burden(medical_area)"
    )
    conn.execute(
        "CREATE INDEX IF NOT EXISTS idx_disease_burden_disease "
        "ON disease_burden(disease_name)"
    )
    conn.commit()


def main():
    conn = sqlite3.connect(DB_PATH)

    # 既存データチェック
    cur = conn.cursor()
    cur.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='disease_burden'"
    )
    if cur.fetchone():
        cur.execute("SELECT COUNT(*) FROM disease_burden")
        existing = cur.fetchone()[0]
        if existing > 0:
            print(f"disease_burdenテーブルに既に{existing}件あります。削除して再作成します。")
            conn.execute("DROP TABLE disease_burden")
            conn.commit()

    create_table(conn)

    total_inserted = 0

    for indicator in INDICATORS:
        filepath = BASE_DIR / indicator["file"]
        if not filepath.exists():
            print(f"  ⚠ ファイル未発見: {indicator['file']} → スキップ")
            continue

        print(f"\n処理中: {indicator['disease']}")
        print(f"  ファイル: {indicator['file']}")
        print(f"  異常基準: {indicator['description']}")

        data = extract_chiba_data(filepath, indicator["abnormal_categories"])

        for area_code, counts in sorted(data.items()):
            area_name = CHIBA_AREAS[area_code]
            abnormal = counts["abnormal"]
            total = counts["total"]
            rate = (abnormal / total * 100) if total > 0 else 0

            conn.execute(
                """INSERT INTO disease_burden
                   (municipality, medical_area, disease_name,
                    patient_count, medical_cost, year, data_source)
                   VALUES (?, ?, ?, ?, NULL, ?, ?)""",
                (
                    area_name,
                    area_name,
                    indicator["disease"],
                    abnormal,
                    YEAR,
                    DATA_SOURCE,
                ),
            )
            total_inserted += 1
            print(f"    {area_name}: {abnormal:,}人 / {total:,}人 ({rate:.1f}%)")

    conn.commit()

    # サマリー表示
    print(f"\n{'='*60}")
    print(f"登録完了: {total_inserted}件")
    print(f"{'='*60}")

    cur.execute("""
        SELECT disease_name, COUNT(*), SUM(patient_count)
        FROM disease_burden
        GROUP BY disease_name
        ORDER BY SUM(patient_count) DESC
    """)
    print("\n【疾患別サマリー】")
    for disease, cnt, total_patients in cur.fetchall():
        print(f"  {disease}: {cnt}圏域, 計{total_patients:,}人")

    cur.execute("""
        SELECT medical_area, COUNT(*), SUM(patient_count)
        FROM disease_burden
        GROUP BY medical_area
        ORDER BY SUM(patient_count) DESC
    """)
    print("\n【医療圏別サマリー（全疾患合計）】")
    for area, cnt, total_patients in cur.fetchall():
        print(f"  {area}: {cnt}指標, 計{total_patients:,}人")

    conn.close()
    print("\n完了！")


if __name__ == "__main__":
    main()
