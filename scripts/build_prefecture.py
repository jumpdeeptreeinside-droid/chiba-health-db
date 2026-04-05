#!/usr/bin/env python3
"""
都道府県DB一括構築スクリプト
都道府県コードを指定するだけで Layer 1,2,3,5 を自動構築する

使い方: python3 build_prefecture.py --code 13  (東京都)

Layer 1: 薬局データ（全国CSVから抽出）
Layer 2: 人口メッシュ（国土数値情報GeoJSON）
Layer 3: NDB特定健診（既存Excelから抽出）
Layer 5: 医療機関（厚生局Excelから抽出 ※要手動DL or 別途対応）
※Layer 4（健康増進計画PDF）はURL調査が必要なため別タスク

ジオコーディング: 座標なし薬局は国土地理院APIで補完
"""

import argparse
import csv
import json
import math
import os
import re
import sqlite3
import subprocess
import sys
import time
import urllib.parse
import urllib.request
import zipfile
import io
from collections import defaultdict
from pathlib import Path

# 都道府県コード→名称
PREF_NAMES = {
    1:"北海道",2:"青森県",3:"岩手県",4:"宮城県",5:"秋田県",6:"山形県",7:"福島県",
    8:"茨城県",9:"栃木県",10:"群馬県",11:"埼玉県",12:"千葉県",13:"東京都",14:"神奈川県",
    15:"新潟県",16:"富山県",17:"石川県",18:"福井県",19:"山梨県",20:"長野県",
    21:"岐阜県",22:"静岡県",23:"愛知県",24:"三重県",
    25:"滋賀県",26:"京都府",27:"大阪府",28:"兵庫県",29:"奈良県",30:"和歌山県",
    31:"鳥取県",32:"島根県",33:"岡山県",34:"広島県",35:"山口県",
    36:"徳島県",37:"香川県",38:"愛媛県",39:"高知県",
    40:"福岡県",41:"佐賀県",42:"長崎県",43:"熊本県",44:"大分県",45:"宮崎県",46:"鹿児島県",47:"沖縄県",
}

CHIBA_PDF_DB = Path.home() / "chiba_pdf_db"
NDB_FILES = [
    ("ndb_hba1c_iryo.xlsx", "糖尿病（HbA1c高値）", ["8.4以上", "8.0以上8.4未満", "6.5以上8.0未満"]),
    ("ndb_bp_iryo.xlsx", "高血圧（収縮期血圧高値）", ["180以上", "160以上180未満", "140以上160未満"]),
    ("ndb_ldl_iryo.xlsx", "脂質異常症（LDL高値）", ["180以上", "160以上180未満", "140以上160未満"]),
    ("ndb_bmi_iryo.xlsx", "肥満（BMI高値）", ["40.0以上", "35.0以上40.0未満", "30.0以上35.0未満", "25.0以上30.0未満"]),
    ("ndb_fbs_iryo.xlsx", "糖尿病（空腹時血糖高値）", ["126以上"]),
]

# ====== Layer 1: 薬局 ======
def build_layer1(conn, pref_code, csv_path):
    print("\n--- Layer 1: 薬局 ---")
    import pandas as pd
    df = pd.read_csv(csv_path, encoding='utf-8-sig', low_memory=False)
    chiba = df[df['都道府県コード'] == pref_code].copy()
    print(f"  {PREF_NAMES[pref_code]}薬局数: {len(chiba)}")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS pharmacies (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gmis_id TEXT, name TEXT NOT NULL, address TEXT,
            lat REAL, lon REAL, city_code INTEGER,
            iryo_ken TEXT, homepage TEXT, data_date TEXT DEFAULT '2025-12-01'
        )
    """)
    conn.execute("DELETE FROM pharmacies")

    result = {
        'gmis_id': chiba['ID'].astype(str),
        'name': chiba['名称'],
        'address': chiba['所在地'],
        'lat': pd.to_numeric(chiba['所在地座標（緯度）'], errors='coerce'),
        'lon': pd.to_numeric(chiba['所在地座標（経度）'], errors='coerce'),
        'city_code': pd.to_numeric(chiba['市区町村コード'], errors='coerce').astype('Int64'),
        'homepage': chiba.get('薬局のホームページアドレス'),
    }
    pdf = pd.DataFrame(result)
    pdf.to_sql('pharmacies', conn, if_exists='append', index=False, method='multi', chunksize=500)
    conn.commit()

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM pharmacies")
    print(f"  登録: {cur.fetchone()[0]}件")


# ====== Layer 2: 人口メッシュ ======
def build_layer2(conn, pref_code):
    print("\n--- Layer 2: 人口メッシュ ---")
    pcode = f"{pref_code:02d}"
    zip_url = f"https://nlftp.mlit.go.jp/ksj/gml/data/m500r6/m500r6-24/500m_mesh_2024_{pcode}_GEOJSON.zip"
    zip_path = Path(f"/tmp/mesh500_{pcode}.zip")
    extract_dir = Path(f"/tmp/mesh500_{pcode}")

    if not extract_dir.exists() or not list(extract_dir.rglob("*.geojson")):
        print(f"  ダウンロード: {zip_url}")
        subprocess.run(["curl", "-sL", zip_url, "-o", str(zip_path)], check=True)
        extract_dir.mkdir(exist_ok=True)
        subprocess.run(["unzip", "-o", str(zip_path), "-d", str(extract_dir)], capture_output=True)

    geojson_files = list(extract_dir.rglob("*.geojson"))
    if not geojson_files:
        print("  ERROR: GeoJSONが見つかりません")
        return
    geojson_path = geojson_files[0]

    with open(geojson_path, encoding="utf-8") as f:
        data = json.load(f)

    conn.execute("DROP TABLE IF EXISTS population_mesh")
    conn.execute("""
        CREATE TABLE population_mesh (
            mesh_code TEXT PRIMARY KEY, lat REAL, lon REAL,
            population INTEGER, municipality TEXT, medical_area TEXT,
            pop_2025 INTEGER, pop_2030 INTEGER, pop_2035 INTEGER,
            pop_2040 INTEGER, pop_2050 INTEGER,
            elderly_65_2025 INTEGER, elderly_75_2025 INTEGER,
            elderly_65_2030 INTEGER, elderly_75_2030 INTEGER,
            elderly_65_2035 INTEGER, elderly_75_2035 INTEGER,
            elderly_65_2040 INTEGER, elderly_75_2040 INTEGER,
            elderly_65_2050 INTEGER, elderly_75_2050 INTEGER
        )
    """)

    records = []
    for feat in data["features"]:
        p = feat["properties"]
        mesh_id = str(p.get("MESH_ID", ""))
        shi_code = str(p.get("SHICODE", ""))
        coords = feat["geometry"]["coordinates"]
        ring = coords[0] if feat["geometry"]["type"] == "Polygon" else coords[0][0]
        clat = round(sum(c[1] for c in ring) / len(ring), 8)
        clon = round(sum(c[0] for c in ring) / len(ring), 8)

        def r(v): return round(v) if v else 0

        records.append((
            mesh_id, clat, clon,
            r(p.get("PTN_2020")), shi_code, None,
            r(p.get("PTN_2025")), r(p.get("PTN_2030")), r(p.get("PTN_2035")),
            r(p.get("PTN_2040")), r(p.get("PTN_2050")),
            r(p.get("PTA_2025")), r(p.get("PTE_2025")),
            r(p.get("PTA_2030")), r(p.get("PTE_2030")),
            r(p.get("PTA_2035")), r(p.get("PTE_2035")),
            r(p.get("PTA_2040")), r(p.get("PTE_2040")),
            r(p.get("PTA_2050")), r(p.get("PTE_2050")),
        ))

    conn.executemany("""
        INSERT INTO population_mesh VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, records)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_mesh_pop ON population_mesh(population)")
    conn.commit()

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*), SUM(population) FROM population_mesh")
    cnt, pop = cur.fetchone()
    print(f"  登録: {cnt:,}メッシュ, 総人口{pop:,.0f}")


# ====== Layer 3: NDB ======
def build_layer3(conn, pref_code):
    print("\n--- Layer 3: NDB特定健診 ---")
    import openpyxl

    pref_name = PREF_NAMES[pref_code]
    conn.execute("DROP TABLE IF EXISTS disease_burden")
    conn.execute("""
        CREATE TABLE disease_burden (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            municipality TEXT, medical_area TEXT, disease_name TEXT,
            patient_count INTEGER, medical_cost REAL,
            year INTEGER, data_source TEXT
        )
    """)

    def safe_int(val):
        if val is None or val in ("‐", "-", "－"):
            return 0
        try:
            return int(val)
        except:
            return 0

    total_inserted = 0
    for filename, disease, abnormal_cats in NDB_FILES:
        filepath = CHIBA_PDF_DB / filename
        if not filepath.exists():
            print(f"  {filename}: 未発見、スキップ")
            continue

        wb = openpyxl.load_workbook(filepath, read_only=True)
        ws = wb[wb.sheetnames[0]]
        in_pref = False

        for row in ws.iter_rows(values_only=True):
            if row[0] == pref_name:
                in_pref = True
            elif row[0] is not None and row[0] != pref_name and in_pref:
                break
            if not in_pref:
                continue

            area_code = row[1]
            area_name = row[2]
            category = row[3]
            if category is None:
                continue

            male_total = safe_int(row[11])
            female_total = safe_int(row[19]) if len(row) > 19 else 0
            count = male_total + female_total

            if area_code is not None and area_name is not None:
                current_area = str(area_name)

            if category in abnormal_cats:
                conn.execute(
                    """INSERT INTO disease_burden
                       (municipality, medical_area, disease_name, patient_count, year, data_source)
                       VALUES (?, ?, ?, ?, 2020, 'NDBオープンデータ第8回')""",
                    (current_area, current_area, disease, count)
                )
                total_inserted += 1

        wb.close()

    conn.commit()
    print(f"  登録: {total_inserted}件")


# ====== ジオコーディング ======
def geocode_missing(conn):
    print("\n--- ジオコーディング ---")
    cur = conn.cursor()
    cur.execute("""
        SELECT id, address FROM pharmacies
        WHERE (lat = 0 OR lon = 0 OR lat IS NULL OR lon IS NULL)
        AND address IS NOT NULL AND address != ''
    """)
    targets = cur.fetchall()
    if not targets:
        print("  対象なし")
        return

    print(f"  対象: {len(targets)}件")
    success = 0
    for i, (pid, address) in enumerate(targets):
        addr = re.sub(r'[（(][^)）]*[)）]', '', address)
        addr = re.sub(r'\d+[FＦ階].*$', '', addr)
        addr = addr.translate(str.maketrans('０１２３４５６７８９', '0123456789')).strip()

        for q in [addr, re.sub(r'\d+[-ー－]\d+.*$', '', addr)]:
            try:
                url = f"https://msearch.gsi.go.jp/address-search/AddressSearch?q={urllib.parse.quote(q)}"
                with urllib.request.urlopen(url, timeout=10) as resp:
                    data = json.loads(resp.read())
                if data:
                    coords = data[0]["geometry"]["coordinates"]
                    cur.execute("UPDATE pharmacies SET lat=?, lon=? WHERE id=?",
                                (coords[1], coords[0], pid))
                    success += 1
                    break
            except:
                pass
        time.sleep(0.5)
        if (i + 1) % 200 == 0:
            conn.commit()
            print(f"  進捗: {i+1}/{len(targets)} (成功={success})")

    conn.commit()
    print(f"  完了: {success}/{len(targets)}")


# ====== メイン ======
def main():
    parser = argparse.ArgumentParser(description="都道府県DB構築")
    parser.add_argument("--code", type=int, required=True, help="都道府県コード (1-47)")
    args = parser.parse_args()

    pref_code = args.code
    if pref_code not in PREF_NAMES:
        print(f"ERROR: 無効な都道府県コード: {pref_code}")
        sys.exit(1)

    pref_name = PREF_NAMES[pref_code]
    pcode = f"{pref_code:02d}"
    db_dir = Path.home() / f"prefdb_{pcode}_{pref_name}"
    db_dir.mkdir(exist_ok=True)
    db_path = db_dir / f"{pcode}_iryo.db"

    print(f"=== {pref_name}（コード{pcode}）DB構築 ===")
    print(f"出力: {db_path}")

    # 全国薬局CSV
    csv_path = CHIBA_PDF_DB / "05_pharmacy_20251201.csv"
    if not csv_path.exists():
        print(f"ERROR: 全国薬局CSV未発見: {csv_path}")
        sys.exit(1)

    conn = sqlite3.connect(db_path)

    build_layer1(conn, pref_code, csv_path)
    build_layer2(conn, pref_code)
    build_layer3(conn, pref_code)
    geocode_missing(conn)

    # 最終サマリー
    cur = conn.cursor()
    print(f"\n{'='*50}")
    print(f"{pref_name} DB構築完了")
    print(f"{'='*50}")
    for tbl in ["pharmacies", "population_mesh", "disease_burden"]:
        cur.execute(f"SELECT COUNT(*) FROM {tbl}")
        print(f"  {tbl}: {cur.fetchone()[0]:,}件")

    cur.execute("SELECT SUM(population) FROM population_mesh")
    print(f"  総人口: {cur.fetchone()[0]:,.0f}")

    conn.close()
    print(f"\n保存: {db_path}")


if __name__ == "__main__":
    main()
