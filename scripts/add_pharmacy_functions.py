#!/usr/bin/env python3
"""
厚生局 施設基準届出データから薬局機能フラグを pharmacies テーブルに追加するスクリプト

データソース: 関東信越厚生局（千葉）/ 近畿厚生局（大阪）
  届出受理医療機関名簿（薬局）
"""

import os
import re
import sqlite3
import zipfile
import io
import openpyxl
from collections import defaultdict
from pathlib import Path

# 届出名称 → フラグカラムのマッピング
FUNCTION_FLAGS = {
    "kakaritsuke": {
        "label": "かかりつけ薬剤師",
        "patterns": ["かかりつけ薬剤師"],
    },
    "chiiki_shien": {
        "label": "地域支援体制加算",
        "patterns": ["地域支援体制加算"],
    },
    "zaitaku_sogo": {
        "label": "在宅薬学総合体制",
        "patterns": ["在宅薬学総合体制"],
    },
    "mukin": {
        "label": "無菌製剤処理",
        "patterns": ["無菌製剤処理"],
    },
    "zaitaku_cv": {
        "label": "在宅中心静脈栄養",
        "patterns": ["在宅中心静脈栄養"],
    },
    "zaitaku_mayaku": {
        "label": "在宅医療用麻薬",
        "patterns": ["在宅患者医療用麻薬"],
    },
    "medical_dx": {
        "label": "医療DX推進",
        "patterns": ["医療ＤＸ推進", "医療DX推進"],
    },
    "generic": {
        "label": "後発医薬品調剤体制",
        "patterns": ["後発医薬品調剤体制"],
    },
    "renkei_kyoka": {
        "label": "連携強化加算",
        "patterns": ["連携強化加算"],
    },
}


def parse_shisetsu_excel(filepath):
    """
    施設基準Excelから薬局番号 → 届出名称リストを抽出
    Returns: {医療機関番号: [届出名称1, 届出名称2, ...]}
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # ヘッダー行を見つける
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if row[0] == "項番":
            header_row = i
            break

    if header_row is None:
        print(f"  WARNING: ヘッダーが見つかりません: {filepath}")
        wb.close()
        return {}

    result = defaultdict(list)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= header_row:
            continue
        kikan_no = str(row[4] or "").strip()
        todokede = str(row[13] or "").strip()
        if kikan_no and todokede:
            result[kikan_no].append(todokede)

    wb.close()
    return dict(result)


def match_pharmacy_to_kikan(cur, kikan_data):
    """
    pharmacies テーブルの gmis_id から医療機関番号を逆引きしてマッチング
    gmis_id は数字列で、下7桁が医療機関番号に対応することが多い
    """
    # 全薬局を取得
    cur.execute("SELECT id, gmis_id, name, address FROM pharmacies")
    pharmacies = cur.fetchall()

    # 医療機関番号のセット
    kikan_set = set(kikan_data.keys())

    matched = {}
    for pid, gmis_id, name, address in pharmacies:
        if not gmis_id:
            continue
        # gmis_id の下7桁
        gid = str(gmis_id).strip()
        candidates = [
            gid[-7:] if len(gid) >= 7 else gid,
            gid[-8:] if len(gid) >= 8 else None,
            gid,
        ]
        for c in candidates:
            if c and c in kikan_set:
                matched[pid] = kikan_data[c]
                break

    return matched


def match_by_name_address(cur, kikan_excel_path):
    """
    名称マッチングでのフォールバック
    """
    wb = openpyxl.load_workbook(kikan_excel_path, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # Excel: {名称: {医療機関番号}}
    name_to_kikan = {}
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == "項番":
            header_found = True
            continue
        if not header_found:
            continue
        name = str(row[7] or "").strip()
        kikan_no = str(row[4] or "").strip()
        if name and kikan_no:
            name_to_kikan[name] = kikan_no
    wb.close()

    # DB薬局
    cur.execute("SELECT id, name FROM pharmacies")
    pharmacies = cur.fetchall()

    # 名称でマッチ
    matched_kikan = {}
    for pid, pname in pharmacies:
        pname_clean = str(pname).strip().replace("　", " ")
        for ename, kno in name_to_kikan.items():
            ename_clean = ename.strip().replace("　", " ")
            if pname_clean == ename_clean:
                matched_kikan[pid] = kno
                break

    return matched_kikan


def process_db(db_path, excel_path, label):
    print(f"\n{'='*50}")
    print(f"薬局機能フラグ追加: {label}")
    print(f"{'='*50}")

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # カラム追加
    for flag_name in FUNCTION_FLAGS:
        try:
            cur.execute(f"ALTER TABLE pharmacies ADD COLUMN func_{flag_name} INTEGER DEFAULT 0")
            print(f"  カラム追加: func_{flag_name}")
        except:
            pass
    # 加算レベル（地域支援体制の等級）
    try:
        cur.execute("ALTER TABLE pharmacies ADD COLUMN chiiki_shien_level INTEGER DEFAULT 0")
    except:
        pass
    conn.commit()

    # Excelパース
    print(f"  Excel解析: {excel_path.name}")
    kikan_data = parse_shisetsu_excel(excel_path)
    print(f"  医療機関数: {len(kikan_data)}")

    # gmis_idでマッチング
    matched = match_pharmacy_to_kikan(cur, kikan_data)
    print(f"  gmis_idマッチ: {len(matched)}件")

    # 名称マッチングで追加
    name_matched = match_by_name_address(cur, excel_path)
    for pid, kno in name_matched.items():
        if pid not in matched and kno in kikan_data:
            matched[pid] = kikan_data[kno]
    print(f"  名称マッチ追加後: {len(matched)}件")

    # フラグ設定
    flag_counts = defaultdict(int)
    for pid, todokede_list in matched.items():
        for todokede in todokede_list:
            for flag_name, config in FUNCTION_FLAGS.items():
                if any(pat in todokede for pat in config["patterns"]):
                    cur.execute(f"UPDATE pharmacies SET func_{flag_name}=1 WHERE id=?", (pid,))
                    flag_counts[flag_name] += 1

            # 地域支援体制の等級
            for level in [4, 3, 2, 1]:
                if f"地域支援体制加算{level}" in todokede or f"地域支援体制加算{level}" in todokede.replace("４","4").replace("３","3").replace("２","2").replace("１","1"):
                    cur.execute("UPDATE pharmacies SET chiiki_shien_level=? WHERE id=? AND chiiki_shien_level<?",
                                (level, pid, level))
                    break

    conn.commit()

    # サマリー
    print(f"\n  機能フラグ集計:")
    cur.execute("SELECT COUNT(*) FROM pharmacies")
    total = cur.fetchone()[0]
    for flag_name, config in FUNCTION_FLAGS.items():
        cur.execute(f"SELECT COUNT(*) FROM pharmacies WHERE func_{flag_name}=1")
        cnt = cur.fetchone()[0]
        rate = cnt / total * 100 if total else 0
        print(f"    {config['label']:<20} {cnt:>5}件 ({rate:.1f}%)")

    # 医療圏別のかかりつけ・地域支援体制
    print(f"\n  医療圏別 かかりつけ率・地域支援体制率:")
    cur.execute("""
        SELECT iryo_ken,
               COUNT(*) as total,
               SUM(func_kakaritsuke) as kk,
               SUM(func_chiiki_shien) as cs,
               SUM(func_mukin) as mk
        FROM pharmacies
        WHERE ds_only = 0 OR ds_only IS NULL
        GROUP BY iryo_ken
        ORDER BY COUNT(*) DESC
    """)
    for area, tot, kk, cs, mk in cur.fetchall():
        kk_r = kk / tot * 100 if tot else 0
        cs_r = cs / tot * 100 if tot else 0
        mk_r = mk / tot * 100 if tot else 0
        print(f"    {area or '不明':<14} かかりつけ{kk_r:>5.1f}% 地域支援{cs_r:>5.1f}% 無菌{mk_r:>4.1f}%")

    conn.close()


if __name__ == "__main__":
    # 千葉
    chiba_db = Path.home() / "chiba_pdf_db" / "chiba_iryo.db"
    chiba_excel = Path.home() / "chiba_pdf_db" / "shisetsu_yakkyoku_r0803" / "12届出受理医療機関名簿（薬局）千葉r0803.xlsx"

    if chiba_db.exists() and chiba_excel.exists():
        process_db(chiba_db, chiba_excel, "千葉県")

    # 大阪: 近畿厚生局からダウンロードが必要
    osaka_db = Path.home() / "osaka_pdf_db" / "osaka_iryo.db"
    osaka_shisetsu_dir = Path.home() / "osaka_pdf_db" / "shisetsu_yakkyoku"

    if osaka_db.exists():
        # 近畿厚生局からDLする処理は別途
        osaka_excels = list(osaka_shisetsu_dir.rglob("*.xlsx")) if osaka_shisetsu_dir.exists() else []
        if osaka_excels:
            process_db(osaka_db, osaka_excels[0], "大阪府")
        else:
            print(f"\n大阪: 施設基準データが見つかりません。")
            print(f"  近畿厚生局からダウンロードが必要:")
            print(f"  https://kouseikyoku.mhlw.go.jp/kinki/chousa/shitei.html")
